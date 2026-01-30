#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ОТЧЕТ ПО НОВЫМ ДАННЫМ ФИКСА (собранным с 5 января 2026)
Данные с collection_date >= 2026-01-05
"""

import sqlite3
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / 'data' / 'fiksa_database.db'

print('\n' + '='*80)
print('📊 ОТЧЕТ ПО НОВЫМ ДАННЫМ ФИКСА (с 5 января 2026)')
print('='*80)

conn = sqlite3.connect(DB_PATH)

# Проверяем данные по collection_date
print('\n[1] Проверка данных по дате сбора...')

collection_info = pd.read_sql_query("""
SELECT 
    collection_date as "Дата сбора",
    COUNT(*) as "Записей",
    MIN(call_date) as "Мин дата звонка",
    MAX(call_date) as "Макс дата звонка",
    COUNT(DISTINCT operator_name) as "Операторов"
FROM fiksa_records
GROUP BY collection_date
ORDER BY collection_date DESC
""", conn)

print(collection_info.to_string(index=False))

# Определяем последнюю дату сбора
latest_collection = collection_info['Дата сбора'].iloc[0]
print(f'\n📅 Последняя дата сбора данных: {latest_collection}')

# Получаем данные с последней даты сбора
print(f'\n[2] Анализ данных с последней даты сбора ({latest_collection})...')

latest_data_stats = pd.read_sql_query(f"""
SELECT 
    COUNT(*) as "Всего записей",
    COUNT(DISTINCT operator_name) as "Операторов",
    COUNT(DISTINCT card_number) as "Уникальных карт",
    SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) as "Положительных",
    SUM(CASE WHEN status LIKE '%Отрицательн%' THEN 1 ELSE 0 END) as "Отрицательных",
    MIN(call_date) as "Самый ранний звонок",
    MAX(call_date) as "Самый поздний звонок"
FROM fiksa_records
WHERE collection_date = '{latest_collection}'
""", conn)

print(latest_data_stats.T.to_string())

# Статистика по операторам с последней даты сбора
print('\n[3] Статистика по операторам...')

operator_stats = pd.read_sql_query(f"""
SELECT 
    operator_name as "Оператор",
    COUNT(*) as "Всего",
    SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) as "Положительных",
    SUM(CASE WHEN status LIKE '%Отрицательн%' THEN 1 ELSE 0 END) as "Отрицательных",
    COUNT(DISTINCT card_number) as "Уникальных карт"
FROM fiksa_records
WHERE collection_date = '{latest_collection}'
GROUP BY operator_name
ORDER BY COUNT(*) DESC
""", conn)

operator_stats['% Положит.'] = (operator_stats['Положительных'] / operator_stats['Всего'] * 100).round(1)

print(f'Операторов в последнем сборе: {len(operator_stats)}')

# Статистика по месяцам звонков
print('\n[4] Распределение по месяцам звонков...')

monthly_dist = pd.read_sql_query(f"""
SELECT 
    strftime('%Y-%m', call_date) as "Месяц",
    COUNT(*) as "Записей",
    SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) as "Положительных"
FROM fiksa_records
WHERE collection_date = '{latest_collection}'
  AND call_date IS NOT NULL
GROUP BY strftime('%Y-%m', call_date)
ORDER BY strftime('%Y-%m', call_date) DESC
LIMIT 12
""", conn)

monthly_dist['% Положит.'] = (monthly_dist['Положительных'] / monthly_dist['Записей'] * 100).round(1)
print(monthly_dist.to_string(index=False))

# Все данные с последней даты сбора
print('\n[5] Выгрузка всех данных с последней даты сбора...')

all_latest_data = pd.read_sql_query(f"""
SELECT 
    id as "ID",
    collection_date as "Дата сбора",
    operator_name as "Оператор",
    card_number as "Номер карты",
    full_name as "ФИО",
    phone as "Телефон",
    address as "Адрес",
    status as "Статус",
    call_date as "Дата звонка",
    notes as "Примечания"
FROM fiksa_records
WHERE collection_date = '{latest_collection}'
ORDER BY call_date DESC, operator_name
""", conn)

conn.close()

# Сохранение в Excel
print('\n💾 Сохранение отчета в Excel...')

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = BASE_DIR / 'reports' / f'ФИКСА_НОВЫЕ_ДАННЫЕ_{latest_collection}_{timestamp}.xlsx'
output_file.parent.mkdir(exist_ok=True)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # 1. Информация о сборах
    collection_info.to_excel(writer, sheet_name='📅 Даты сбора', index=False)
    
    # 2. Статистика последнего сбора
    latest_data_stats.T.to_excel(writer, sheet_name='📊 Общая статистика')
    
    # 3. По операторам
    operator_stats.to_excel(writer, sheet_name='👥 По операторам', index=False)
    
    # 4. По месяцам
    monthly_dist.to_excel(writer, sheet_name='📆 По месяцам', index=False)
    
    # 5. Все данные
    all_latest_data.to_excel(writer, sheet_name='📄 Все данные', index=False)

# Форматирование
print('🎨 Применение форматирования...')
wb = load_workbook(output_file)

header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Форматирование заголовков
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Автоширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'

wb.save(output_file)

print('\n' + '='*80)
print('✅ ОТЧЕТ ПО НОВЫМ ДАННЫМ СОЗДАН')
print('='*80)
print(f'\n📁 Файл: {output_file.name}')
print(f'📂 Путь: {output_file}')
print(f'\n📊 СОДЕРЖАНИЕ:')
print(f'  1. Даты сбора - история сборов данных')
print(f'  2. Общая статистика - сводка по последнему сбору')
print(f'  3. По операторам - {len(operator_stats)} операторов')
print(f'  4. По месяцам - распределение звонков')
print(f'  5. Все данные - {len(all_latest_data):,} записей')

print(f'\n📈 КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ (последний сбор {latest_collection}):')
print(f'  • Всего записей: {latest_data_stats["Всего записей"].iloc[0]:,}')
print(f'  • Операторов: {latest_data_stats["Операторов"].iloc[0]}')
print(f'  • Уникальных карт: {latest_data_stats["Уникальных карт"].iloc[0]:,}')
print(f'  • Положительных: {latest_data_stats["Положительных"].iloc[0]:,} ({latest_data_stats["Положительных"].iloc[0]/latest_data_stats["Всего записей"].iloc[0]*100:.1f}%)')
print(f'  • Период звонков: {latest_data_stats["Самый ранний звонок"].iloc[0]} - {latest_data_stats["Самый поздний звонок"].iloc[0]}')

print('\n' + '='*80)
