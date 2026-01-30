#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ОТЧЕТ ПО ДАННЫМ ФИКСА - ПОЛНЫЙ АНАЛИЗ
Все записи из базы данных с детальной статистикой
"""

import sqlite3
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / 'data' / 'fiksa_database.db'

print('\n' + '='*80)
print('📊 ПОЛНЫЙ ОТЧЕТ ПО ДАННЫМ ФИКСА')
print('='*80)

conn = sqlite3.connect(DB_PATH)

# ============================================================================
# 1. ОБЩАЯ СТАТИСТИКА
# ============================================================================
print('\n[1/7] Сбор общей статистики...')

general_stats = pd.read_sql_query("""
SELECT 
    COUNT(*) as "Всего записей",
    COUNT(DISTINCT operator_name) as "Операторов",
    COUNT(DISTINCT card_number) as "Уникальных карт",
    SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) as "Положительных",
    SUM(CASE WHEN status LIKE '%Отрицательн%' THEN 1 ELSE 0 END) as "Отрицательных",
    SUM(CASE WHEN status LIKE '%Не отвечает%' THEN 1 ELSE 0 END) as "Не отвечает",
    MIN(call_date) as "Первая дата",
    MAX(call_date) as "Последняя дата"
FROM fiksa_records
""", conn)

print(general_stats.T.to_string())

# ============================================================================
# 2. СТАТИСТИКА ПО ОПЕРАТОРАМ
# ============================================================================
print('\n[2/7] Анализ по операторам...')

operator_stats = pd.read_sql_query("""
SELECT 
    operator_name as "Оператор",
    COUNT(*) as "Всего",
    SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) as "Положительных",
    SUM(CASE WHEN status LIKE '%Отрицательн%' THEN 1 ELSE 0 END) as "Отрицательных",
    SUM(CASE WHEN status LIKE '%Не отвечает%' THEN 1 ELSE 0 END) as "Не отвечает",
    COUNT(DISTINCT card_number) as "Уникальных карт",
    MIN(call_date) as "Первый звонок",
    MAX(call_date) as "Последний звонок"
FROM fiksa_records
GROUP BY operator_name
ORDER BY COUNT(*) DESC
""", conn)

operator_stats['% Положит.'] = (operator_stats['Положительных'] / operator_stats['Всего'] * 100).round(1)
operator_stats['% Отрицат.'] = (operator_stats['Отрицательных'] / operator_stats['Всего'] * 100).round(1)

print(f'Операторов: {len(operator_stats)}')

# ============================================================================
# 3. СТАТИСТИКА ПО ДАТАМ
# ============================================================================
print('\n[3/7] Анализ по датам...')

daily_stats = pd.read_sql_query("""
SELECT 
    DATE(call_date) as "Дата",
    COUNT(*) as "Всего",
    SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) as "Положительных",
    SUM(CASE WHEN status LIKE '%Отрицательн%' THEN 1 ELSE 0 END) as "Отрицательных",
    COUNT(DISTINCT operator_name) as "Операторов работало"
FROM fiksa_records
WHERE call_date IS NOT NULL
GROUP BY DATE(call_date)
ORDER BY DATE(call_date) DESC
LIMIT 100
""", conn)

daily_stats['% Положит.'] = (daily_stats['Положительных'] / daily_stats['Всего'] * 100).round(1)

# ============================================================================
# 4. СТАТИСТИКА ПО МЕСЯЦАМ
# ============================================================================
print('\n[4/7] Анализ по месяцам...')

monthly_stats = pd.read_sql_query("""
SELECT 
    strftime('%Y-%m', call_date) as "Месяц",
    COUNT(*) as "Всего",
    SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) as "Положительных",
    SUM(CASE WHEN status LIKE '%Отрицательн%' THEN 1 ELSE 0 END) as "Отрицательных",
    COUNT(DISTINCT operator_name) as "Операторов",
    COUNT(DISTINCT card_number) as "Уникальных карт"
FROM fiksa_records
WHERE call_date IS NOT NULL
GROUP BY strftime('%Y-%m', call_date)
ORDER BY strftime('%Y-%m', call_date) DESC
""", conn)

monthly_stats['% Положит.'] = (monthly_stats['Положительных'] / monthly_stats['Всего'] * 100).round(1)

# ============================================================================
# 5. СТАТИСТИКА ПО СТАТУСАМ
# ============================================================================
print('\n[5/7] Анализ по статусам...')

status_stats = pd.read_sql_query("""
SELECT 
    status as "Статус",
    COUNT(*) as "Количество",
    COUNT(DISTINCT operator_name) as "Операторов"
FROM fiksa_records
GROUP BY status
ORDER BY COUNT(*) DESC
""", conn)

status_stats['%'] = (status_stats['Количество'] / status_stats['Количество'].sum() * 100).round(2)

# ============================================================================
# 6. ТОП ОПЕРАТОРЫ ПО РЕЗУЛЬТАТИВНОСТИ
# ============================================================================
print('\n[6/7] ТОП операторов...')

top_operators = pd.read_sql_query("""
SELECT 
    operator_name as "Оператор",
    COUNT(*) as "Всего звонков",
    SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) as "Положительных",
    ROUND(100.0 * SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) / COUNT(*), 1) as "% Положит."
FROM fiksa_records
GROUP BY operator_name
HAVING COUNT(*) >= 100
ORDER BY 
    ROUND(100.0 * SUM(CASE WHEN status LIKE '%Положительн%' THEN 1 ELSE 0 END) / COUNT(*), 1) DESC,
    COUNT(*) DESC
LIMIT 20
""", conn)

# ============================================================================
# 7. ВСЕ ДАННЫЕ (ДЛЯ ДЕТАЛЬНОГО АНАЛИЗА)
# ============================================================================
print('\n[7/7] Выгрузка всех данных...')

all_data = pd.read_sql_query("""
SELECT 
    id as "ID",
    collection_date as "Дата сбора",
    operator_name as "Оператор",
    card_number as "Номер карты",
    full_name as "ФИО",
    phone as "Телефон",
    address as "Адрес",
    status as "Статус",
    call_date as "Дата звонка",
    notes as "Примечания"
FROM fiksa_records
ORDER BY call_date DESC, operator_name
""", conn)

conn.close()

# ============================================================================
# СОХРАНЕНИЕ В EXCEL
# ============================================================================
print('\n💾 Сохранение отчета в Excel...')

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = BASE_DIR / 'reports' / f'ФИКСА_ПОЛНЫЙ_ОТЧЕТ_{timestamp}.xlsx'
output_file.parent.mkdir(exist_ok=True)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # 1. Общая статистика
    general_stats.T.to_excel(writer, sheet_name='📊 Общая статистика')
    
    # 2. По операторам
    operator_stats.to_excel(writer, sheet_name='👥 По операторам', index=False)
    
    # 3. По дням (последние 100)
    daily_stats.to_excel(writer, sheet_name='📅 По дням', index=False)
    
    # 4. По месяцам
    monthly_stats.to_excel(writer, sheet_name='📆 По месяцам', index=False)
    
    # 5. По статусам
    status_stats.to_excel(writer, sheet_name='📋 По статусам', index=False)
    
    # 6. ТОП операторы
    top_operators.to_excel(writer, sheet_name='🏆 ТОП операторы', index=False)
    
    # 7. Все данные
    all_data.to_excel(writer, sheet_name='📄 Все данные', index=False)

# Форматирование
print('🎨 Применение форматирования...')
wb = load_workbook(output_file)

# Стили
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Форматирование заголовков
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Автоширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Закрепление первой строки
    ws.freeze_panes = 'A2'

wb.save(output_file)

print('\n' + '='*80)
print('✅ ОТЧЕТ УСПЕШНО СОЗДАН')
print('='*80)
print(f'\n📁 Файл: {output_file.name}')
print(f'📂 Путь: {output_file}')
print(f'\n📊 СОДЕРЖАНИЕ ОТЧЕТА:')
print(f'  1. Общая статистика - сводка по всей базе')
print(f'  2. По операторам - {len(operator_stats)} операторов')
print(f'  3. По дням - последние {len(daily_stats)} дней')
print(f'  4. По месяцам - {len(monthly_stats)} месяцев')
print(f'  5. По статусам - {len(status_stats)} статусов')
print(f'  6. ТОП операторы - топ 20 по результативности')
print(f'  7. Все данные - {len(all_data):,} записей')

print(f'\n📈 КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ:')
print(f'  • Всего записей: {general_stats["Всего записей"].iloc[0]:,}')
print(f'  • Операторов: {general_stats["Операторов"].iloc[0]}')
print(f'  • Положительных: {general_stats["Положительных"].iloc[0]:,} ({general_stats["Положительных"].iloc[0]/general_stats["Всего записей"].iloc[0]*100:.1f}%)')
print(f'  • Период: {general_stats["Первая дата"].iloc[0]} - {general_stats["Последняя дата"].iloc[0]}')

print('\n' + '='*80)
