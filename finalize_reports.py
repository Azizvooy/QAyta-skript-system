#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Финальная обработка отчетов согласно всем требованиям:
1. Удалить колонку Дата_112
2. Удалить колонки C-E, J-L, R-S, X-Y (после удаления Дата_112)
3. Сортировка по Қўнғироқ қабул қилинган вақт (от ранних к поздним)
4. Нумерация после сортировки
5. В Отрицательные_и_жалобы только со статусом "отриц" (без "заявка закрыта")
6. Удалить префиксы 1./2./3./4. из жалоб
7. Жалобы_по_регионам: добавить Положительные и Не дозвонились
8. Регионы_и_жалобы: pivot (регионы слева, жалобы сверху)
9. Не_найденные_заявки: только не учтенные в основных категориях
10. Границы, автофильтр, ширины колонок
"""

import os
import glob
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment


# Колонки для удаления по НАЗВАНИЯМ (точные имена колонок которые нужно удалить)
# После удаления Дата_112, удаляем эти колонки:
# C-E: Қўнғироқ давомийлиги, Бригадага узатилган вақт, Қўнғироқ якунланган вақт
# J-L: Статус_112, Қўнғироқ қилувчи Ф.И.Ш, (еще одна)
# R-S: Телефон_112, (еще одна)  
# X-Y: (последние)
# НО сохраняем: Регион_112, Служба_112, Телефон_нормализованный, Инцидент_112, Жалоба, Статус_связи
DROP_COLUMN_NAMES = [
    'Қўнғироқ давомийлиги',
    'Бригадага узатилган вақт', 
    'Қўнғироқ якунланган вақт',
    'Статус_112',
    'Қўнғироқ қилувчи Ф.И.Ш',
    'Телефон_112',
    'Ўзи рад этган',
    'Есть_жалоба'
]


def classify_status(val):
    if pd.isna(val):
        return 'Неизвестно'
    text = str(val).strip().lower()
    if not text:
        return 'Неизвестно'
    if 'полож' in text:
        return 'Положительные'
    if 'отриц' in text:
        return 'Отрицательные'
    if any(k in text for k in ['не удалось', 'недозвон', 'не дозвон', 'не ответ', 'занят', 'занято', 'сброс']):
        return 'Не дозвонились'
    return 'Прочее'


def drop_extra_columns(df):
    """Удаляет лишние колонки по именам"""
    return df.drop(columns=DROP_COLUMN_NAMES, errors='ignore')


def clean_complaint_prefix(val):
    """Удаляет префикс 1./2./3./4. из начала жалобы"""
    if pd.isna(val):
        return val
    text = str(val)
    return re.sub(r'^\s*[1-4]\.\s*', '', text)


def add_numbering(df):
    """Добавляет колонку № в начало"""
    df_copy = df.copy()
    if '№' in df_copy.columns:
        df_copy = df_copy.drop(columns=['№'])
    df_copy.insert(0, '№', range(1, len(df_copy) + 1))
    return df_copy


def apply_borders_and_width(ws):
    """Применяет границы и автоширину к листу"""
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)


def replace_sheet(wb, sheet_name, df):
    """Заменяет лист в workbook"""
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name)
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    apply_borders_and_width(ws)
    
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def process_file(path):
    """Обрабатывает один файл согласно всем требованиям"""
    wb = load_workbook(path)
    
    # 1. Обработка листа Детальные
    if 'Детальные' not in wb.sheetnames:
        return
    
    df_det = pd.read_excel(path, sheet_name='Детальные')
    
    # Удаляем Дата_112
    if 'Дата_112' in df_det.columns:
        df_det = df_det.drop(columns=['Дата_112'])
    
    # Очищаем префиксы в жалобах
    if 'Жалоба' in df_det.columns:
        df_det['Жалоба'] = df_det['Жалоба'].apply(clean_complaint_prefix)
    
    # Сортируем по дате
    date_col = 'Қўнғироқ қабул қилинган вақт'
    if date_col in df_det.columns:
        df_det = df_det.sort_values(date_col, na_position='last')
    
    # Удаляем лишние колонки
    df_det = drop_extra_columns(df_det)
    
    # Добавляем нумерацию
    df_det = add_numbering(df_det)
    
    # Сохраняем
    replace_sheet(wb, 'Детальные', df_det)
    
    # 2. Обработка Отрицательные_и_жалобы
    if 'Статус_связи' in df_det.columns:
        status = df_det['Статус_связи'].astype(str)
        neg_mask = status.str.contains('отриц', case=False, na=False)
        closed_mask = status.str.contains('заявка закрыта', case=False, na=False)
        df_neg = df_det[neg_mask & ~closed_mask].copy()
        
        # Убираем колонку № перед удалением позиций
        if '№' in df_neg.columns:
            df_neg = df_neg.drop(columns=['№'])
        
        # Снова применяем удаление колонок и нумерацию
        df_neg = add_numbering(df_neg)
        replace_sheet(wb, 'Отрицательные_и_жалобы', df_neg)
    
    # 3. Обработка Регионы_и_жалобы - делаем pivot
    if 'Регионы_и_жалобы' in wb.sheetnames:
        df_long = pd.read_excel(path, sheet_name='Регионы_и_жалобы')
        
        if {'Регион_112', 'Жалоба', 'Количество'}.issubset(df_long.columns):
            # Очищаем префиксы в названиях жалоб
            df_long['Жалоба'] = df_long['Жалоба'].apply(clean_complaint_prefix)
            
            # Создаем pivot
            pivot = df_long.pivot_table(
                index='Регион_112',
                columns='Жалоба',
                values='Количество',
                aggfunc='sum',
                fill_value=0
            )
            pivot = pivot.reset_index()
            pivot = add_numbering(pivot)
            replace_sheet(wb, 'Регионы_и_жалобы', pivot)
    
    # 4. Обработка Жалобы_по_регионам - добавляем Положительные и Не дозвонились
    region_col = 'Регион_112' if 'Регион_112' in df_det.columns else df_det.columns[1]
    status_col = 'Статус_связи' if 'Статус_связи' in df_det.columns else None
    complaints_col = 'Есть_жалоба' if 'Есть_жалоба' in df_det.columns else None
    
    if complaints_col is None and 'Жалоба' in df_det.columns:
        complaints_series = df_det['Жалоба'].notna() & (df_det['Жалоба'].astype(str).str.strip() != '')
    else:
        complaints_series = df_det[complaints_col].fillna(False) if complaints_col else pd.Series([False]*len(df_det))
    
    statuses = df_det[status_col].apply(classify_status) if status_col else pd.Series(['Неизвестно']*len(df_det))
    
    total = df_det.groupby(region_col).size().rename('Всего')
    complaints = df_det[complaints_series].groupby(region_col).size().rename('Количество_жалоб')
    positives = df_det[statuses.eq('Положительные')].groupby(region_col).size().rename('Положительные')
    noanswer = df_det[statuses.eq('Не дозвонились')].groupby(region_col).size().rename('Не дозвонились')
    
    df_reg = pd.concat([complaints, positives, noanswer, total], axis=1).fillna(0).astype(int).reset_index()
    df_reg = df_reg.rename(columns={region_col: 'Регион_112'})
    df_reg = add_numbering(df_reg)
    replace_sheet(wb, 'Жалобы_по_регионам', df_reg)
    
    # 5. Обработка Не_найденные_заявки
    accounted = complaints_series | statuses.eq('Положительные') | statuses.eq('Не дозвонились') | statuses.eq('Отрицательные')
    missing_df = df_det[~accounted].copy()
    
    if len(missing_df) > 0:
        # Убираем нумерацию
        if '№' in missing_df.columns:
            missing_df = missing_df.drop(columns=['№'])
        missing_df = add_numbering(missing_df)
        replace_sheet(wb, 'Не_найденные_заявки', missing_df)
    elif 'Не_найденные_заявки' in wb.sheetnames:
        wb.remove(wb['Не_найденные_заявки'])
    
    # Сохраняем файл
    wb.save(path)


def main():
    import sys
    
    if len(sys.argv) > 1:
        folder = sys.argv[1]
    else:
        folder = 'reports/Службы Обратная связь за Январь месяц 2026 год'
    
    files = glob.glob(os.path.join(folder, '*.xlsx'))
    count = 0
    
    for path in files:
        if 'ОТРИЦАТЕЛЬНЫЕ' in path or '~$' in path:
            continue
        
        try:
            print(f"Обработка: {os.path.basename(path)}")
            process_file(path)
            count += 1
        except Exception as e:
            print(f"  ⚠️ Ошибка: {e}")
    
    print(f"\n✅ Обработано {count} файлов")
    
    # Обновляем ТОЛЬКО_ОТРИЦАТЕЛЬНЫЕ
    neg_folder = os.path.join(folder, 'ТОЛЬКО_ОТРИЦАТЕЛЬНЫЕ')
    if os.path.exists(neg_folder):
        print("\nОбновление ТОЛЬКО_ОТРИЦАТЕЛЬНЫЕ...")
        for path in files:
            if 'ИНЦИДЕНТЫ' not in path or 'ОТРИЦАТЕЛЬНЫЕ' in path or '~$' in path:
                continue
            
            if 'Отрицательные_и_жалобы' not in pd.ExcelFile(path).sheet_names:
                continue
            
            df = pd.read_excel(path, sheet_name='Отрицательные_и_жалобы')
            out_path = os.path.join(neg_folder, os.path.basename(path).replace('ИНЦИДЕНТЫ', 'ОТРИЦАТЕЛЬНЫЕ'))
            
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)
            replace_sheet(wb, 'Отрицательные_и_жалобы', df)
            wb.save(out_path)
            print(f"  ✓ {os.path.basename(out_path)}")
        
        print("✅ Обновлены файлы ТОЛЬКО_ОТРИЦАТЕЛЬНЫЕ")


if __name__ == '__main__':
    main()
