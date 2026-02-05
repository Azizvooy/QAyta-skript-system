#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
===========================================================================
ОБНОВЛЕННАЯ ОБРАБОТКА ДАННЫХ И ГЕНЕРАЦИЯ ОТЧЕТОВ
===========================================================================
Использует правильную логику обработки данных из последней версии
===========================================================================
"""

import sqlite3
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / 'data' / 'fiksa_database.db'
EXPORT_DIR = BASE_DIR / 'exported_sheets'
OUTPUT_DIR = BASE_DIR / 'output' / 'reports'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print('\n' + '='*80)
print('📊 ОБНОВЛЕННАЯ ГЕНЕРАЦИЯ ОТЧЕТОВ')
print('='*80)

def load_exported_data():
    """Загрузка данных из экспортированных CSV"""
    print('\n[1/5] Загрузка экспортированных данных...')
    
    all_data = []
    csv_files = list(EXPORT_DIR.rglob('*.csv'))
    
    print(f'  Найдено CSV файлов: {len(csv_files)}')
    
    for csv_file in csv_files:
        try:
            df = pd.read_csv(csv_file, encoding='utf-8-sig', low_memory=False)
            
            if len(df) == 0:
                continue
            
            # Нормализуем колонки
            normalized = {}
            
            # Номер карты
            for col in ['Номер карты', 'Код карты', 'Кодкарты', '№']:
                if col in df.columns:
                    normalized['card_number'] = df[col]
                    break
            
            # Статус
            for col in ['Статус связи', 'Причина/Статус', 'Статус']:
                if col in df.columns:
                    normalized['status'] = df[col]
                    break
            
            # Оператор
            for col in ['Оператор', 'USER', 'Пользователь']:
                if col in df.columns:
                    normalized['operator_name'] = df[col]
                    break
                    
            # Дата
            for col in ['Дата фиксации', 'Время фиксации', 'Дата открытия карты', 'Дата']:
                if col in df.columns:
                    normalized['call_date'] = df[col]
                    break
            
            # Телефон
            for col in ['Телефон', 'Номер телефона', 'Phone']:
                if col in df.columns:
                    normalized['phone'] = df[col]
                    break
            
            # ФИО
            for col in ['ФИО', 'Полное имя', 'Имя']:
                if col in df.columns:
                    normalized['full_name'] = df[col]
                    break
            
            # Адрес
            for col in ['Адрес', 'Address']:
                if col in df.columns:
                    normalized['address'] = df[col]
                    break
            
            if normalized:
                df_normalized = pd.DataFrame(normalized)
                df_normalized['source'] = csv_file.name
                all_data.append(df_normalized)
                
        except Exception as e:
            print(f'  ⚠️ Ошибка в {csv_file.name}: {str(e)[:50]}')
    
    if all_data:
        combined = pd.concat(all_data, ignore_index=True)
        print(f'  ✅ Загружено записей: {len(combined):,}')
        return combined
    
    return pd.DataFrame()

def load_database_data():
    """Загрузка данных из базы данных"""
    print('\n[2/5] Загрузка данных из базы данных...')
    
    try:
        conn = sqlite3.connect(DB_PATH)
        
        # Объединенные данные фикса + заявки
        query_full = '''
            SELECT 
                fr.id,
                fr.collection_date,
                fr.call_date,
                DATE(fr.call_date) as call_day,
                fr.operator_name,
                fr.card_number,
                COALESCE(fr.phone, ch.caller_phone) as phone,
                fr.full_name,
                fr.address,
                fr.status,
                ch.service_name,
                ch.status as application_status,
                ch.region,
                ch.district,
                ch.incident_number,
                ch.reason,
                ch.description,
                fr.notes
            FROM fiksa_records fr
            LEFT JOIN call_history_112 ch ON fr.card_number = ch.card_number
            WHERE fr.operator_name IS NOT NULL
            ORDER BY fr.call_date DESC, fr.operator_name
        '''
        
        df = pd.read_sql_query(query_full, conn)
        conn.close()
        
        print(f'  ✅ Загружено записей из БД: {len(df):,}')
        return df
        
    except Exception as e:
        print(f'  ⚠️ Ошибка загрузки из БД: {e}')
        return pd.DataFrame()

def generate_daily_report(df):
    """Отчет по дням"""
    print('\n[3/5] Генерация отчета по дням...')
    
    # Обработка дат с учетом разных форматов
    df['call_day'] = pd.to_datetime(df['call_date'], format='mixed', dayfirst=True, errors='coerce').dt.date
    
    # Для count используем любую существующую колонку
    count_col = 'card_number' if 'card_number' in df.columns else 'status'
    
    daily = df.groupby('call_day').agg({
        count_col: 'count',
        'operator_name': 'nunique',
        'status': lambda x: (x.str.contains('Положительн', case=False, na=False)).sum(),
    }).rename(columns={
        count_col: 'Всего фиксаций',
        'operator_name': 'Операторов',
        'status': 'Положительных'
    })
    
    daily['% Положит.'] = round(daily['Положительных'] / daily['Всего фиксаций'] * 100, 1)
    daily = daily.sort_index(ascending=False).head(60)
    
    print(f'  ✅ Сгенерировано дней: {len(daily)}')
    return daily.reset_index()

def generate_operator_rating(df):
    """Рейтинг операторов"""
    print('\n[4/5] Генерация рейтинга операторов...')
    
    def count_positive(x):
        return (x.str.contains('Положительн', case=False, na=False)).sum()
    
    def count_negative(x):
        return (x.str.contains('Отрицательн', case=False, na=False)).sum()
    
    # Для count используем любую существующую колонку
    count_col = 'card_number' if 'card_number' in df.columns else 'status'
    
    rating = df.groupby('operator_name').agg({
        count_col: 'count',
        'status': [count_positive, count_negative]
    })
    
    rating.columns = ['Всего фиксаций', 'Положительных', 'Отрицательных']
    rating['% Положит.'] = round(rating['Положительных'] / rating['Всего фиксаций'] * 100, 1)
    rating = rating.sort_values('Положительных', ascending=False)
    
    print(f'  ✅ Операторов в рейтинге: {len(rating)}')
    return rating.reset_index()

def save_excel_report(df_full, df_daily, df_rating):
    """Сохранение Excel отчета"""
    print('\n[5/5] Сохранение Excel отчета...')
    
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    output_file = OUTPUT_DIR / f'📊_ПОЛНЫЙ_ОТЧЕТ_{timestamp}.xlsx'
    
    # Ограничиваем последние 100,000 записей для Excel (лимит 1,048,576)
    df_recent = df_full.tail(100000)
    
    print(f'  💾 Сохранение {len(df_recent):,} из {len(df_full):,} записей (последние 100K)')
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_recent.to_excel(writer, sheet_name='Все данные (100K)', index=False)
        df_daily.to_excel(writer, sheet_name='По дням', index=False)
        df_rating.to_excel(writer, sheet_name='Рейтинг операторов', index=False)
    
    # Форматирование
    wb = load_workbook(output_file)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Авто-ширина
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    
    file_size = output_file.stat().st_size / (1024 * 1024)
    print(f'  ✅ Отчет сохранен: {output_file.name}')
    print(f'  📦 Размер: {file_size:.2f} МБ')
    
    return output_file

def main():
    """Основная функция"""
    # Загрузка данных
    df_export = load_exported_data()
    df_db = load_database_data()
    
    # Используем данные из БД если есть, иначе из экспорта
    if not df_db.empty:
        df_main = df_db
    elif not df_export.empty:
        df_main = df_export
    else:
        print('\n❌ Нет данных для обработки!')
        return
    
    # Генерация отчетов
    df_daily = generate_daily_report(df_main)
    df_rating = generate_operator_rating(df_main)
    
    # Сохранение
    output_file = save_excel_report(df_main, df_daily, df_rating)
    
    print('\n' + '='*80)
    print('✅ ОТЧЕТ УСПЕШНО СОЗДАН!')
    print('='*80)
    print(f'\n📊 Статистика:')
    print(f'   Всего записей: {len(df_main):,}')
    print(f'   Операторов: {df_main["operator_name"].nunique()}')
    print(f'   Период: {len(df_daily)} дней')
    print(f'\n📁 Файл: {output_file}')
    print('='*80)

if __name__ == '__main__':
    main()
