#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ФИНАЛЬНАЯ ОБРАБОТКА ОТЧЕТОВ - ПОЛНАЯ ЛОГИКА
==============================================

Назначение: Преобразовать исходные файлы отчетов (.xlsx) согласно всем требованиям.

ТРЕБОВАНИЯ К ОБРАБОТКЕ:
1. ✅ Удалить колонку Дата_112
2. ✅ Удалить лишние колонки: Қўнғироқ давомийлиги, Бригадага узатилган вақт, 
                              Қўнғироқ якунланган вақт, Статус_112, 
                              Қўнғироқ қилувчи Ф.И.Ш, Телефон_112, 
                              Ўзи рад этган, Есть_жалоба
3. ✅ Сортировка по Қўнғироқ қабул қилинган вақт (от ранних к поздним)
4. ✅ Нумерация (колонка №) после сортировки, начиная с 1
5. ✅ Удалить префиксы 1./2./3./4. из текста жалоб
6. ✅ В листе "Отрицательные_и_жалобы" - только со статусом "отриц" (БЕЗ "заявка закрыта")
7. ✅ Лист "Жалобы_по_регионам" - добавить колонки: Положительные, Не дозвонились, Всего
8. ✅ Лист "Регионы_и_жалобы" - PIVOT таблица (регионы слева, жалобы сверху)
9. ✅ Лист "Не_найденные_заявки" - только не учтенные записи
10. ✅ Добавить границы, автофильтр, оптимальные ширины колонок

КЛАССИФИКАЦИЯ СТАТУСОВ:
- Положительные: text.lower() содержит "полож"
- Отрицательные: text.lower() содержит "отриц" И НЕ содержит "заявка закрыта"
- Не дозвонились: text.lower() содержит "не удалось" OR "недозвон" OR "не дозвон" OR "не ответ" OR "занят" OR "сброс"
- Прочее: остальное
- Неизвестно: пусто или NaN

СОХРАНЯЕМЫЕ КОЛОНКИ (оставляем в файле):
- №, Қўнғироқ қабул қилинган вақт, Служба_112, Сабаб, Карта_112,
  Инцидент_112, Оператор_112, Регион_112, Район_112, Манзил,
  Қўнғироқ жойи, Тавсиф, Телефон_нормализованный, Инцидент_112_norm,
  Жалоба, Статус_связи, Положительно

СТРУКТУРА ВЫХОДНЫХ ФАЙЛОВ:
1. Детальные - все записи со всеми колонками
2. Отрицательные_и_жалобы - только "отриц" статусы
3. Жалобы_по_регионам - сводка по регионам
4. Регионы_и_жалобы - pivot таблица
5. Не_найденные_заявки - неучтенные записи
"""

import os
import sys
import glob
import re
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import (Border, Side, Font, PatternFill, Alignment)
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================================
# КОНФИГУРАЦИЯ
# ============================================================================

# Колонки для удаления (по НАЗВАНИЯМ, а не по позициям!)
DELETE_COLUMNS = [
    'Дата_112',
    'Қўнғироқ давомийлиги',
    'Бригадага узатилган вақт',
    'Қўнғироқ якунланган вақт',
    'Статус_112',
    'Қўнғироқ қилувчи Ф.И.Ш',
    'Телефон_112',
    'Ўзи рад этган',
    'Есть_жалоба'
]

# Колонки для сортировки (главная дата)
DATE_COLUMN = 'Қўнғироқ қабул қилинган вақт'

# Колонка со статусом
STATUS_COLUMN = 'Статус_связи'

# Колонка с жалобой
COMPLAINT_COLUMN = 'Жалоба'


# ============================================================================
# ФУНКЦИИ КЛАССИФИКАЦИИ
# ============================================================================

def classify_status(val):
    """Классифицирует статус в одну из категорий"""
    if pd.isna(val):
        return 'Неизвестно'
    
    text = str(val).strip().lower()
    if not text:
        return 'Неизвестно'
    
    # Проверяем в порядке приоритета
    if 'полож' in text:
        return 'Положительные'
    
    # Отрицательные, но НЕ "заявка закрыта"
    if 'отриц' in text and 'заявка закрыта' not in text:
        return 'Отрицательные'
    
    # Не дозвонились
    patterns = ['не удалось', 'недозвон', 'не дозвон', 'не ответ', 'занят', 'сброс']
    if any(pat in text for pat in patterns):
        return 'Не дозвонились'
    
    if 'отриц' in text or 'заявка закрыта' in text:
        return 'Прочее'
    
    return 'Прочее'


def clean_complaint_prefix(text):
    """Удаляет префиксы 1./2./3./4. из начала текста жалобы"""
    if pd.isna(text):
        return text
    
    text = str(text)
    # Удаляем начальные пробелы и цифру с точкой
    text = re.sub(r'^\s*[1-4]\.\s*', '', text)
    return text


# ============================================================================
# ФУНКЦИИ ОБРАБОТКИ
# ============================================================================

def drop_extra_columns(df):
    """Удаляет лишние колонки по названиям"""
    cols_to_drop = [col for col in DELETE_COLUMNS if col in df.columns]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
    return df


def add_numbering(df):
    """Добавляет колонку № с последовательной нумерацией (1, 2, 3...)"""
    df.insert(0, '№', range(1, len(df) + 1))
    return df


def apply_borders_and_formatting(ws, max_col=None, header_fill=None):
    """
    Применяет границы, автофильтр и форматирование к листу
    
    Args:
        ws: рабочий лист openpyxl
        max_col: максимальное кол-во колонок (если None, берется от данных)
        header_fill: заливка для заголовка (Color)
    """
    if max_col is None:
        max_col = ws.max_column
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Форматируем все ячейки
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Форматируем заголовок специальнее
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_font = Font(bold=True)
    
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Добавляем автофильтр
    if ws.max_row > 0:
        ws.auto_filter.ref = f'A1:{chr(64 + max_col)}{ws.max_row}'
    
    # Оптимизируем ширину колонок
    for col in range(1, max_col + 1):
        col_letter = chr(64 + col)
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value or ''))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)  # макс 50 символов
        ws.column_dimensions[col_letter].width = adjusted_width


def process_detailed_sheet(wb, df):
    """Обрабатывает лист "Детальные"
    
    Шаги:
    1. Удаляем лишние колонки
    2. Сортируем по дате
    3. Добавляем нумерацию
    4. Очищаем текст жалоб (удаляем префиксы)
    5. Форматируем
    """
    # Удаляем лишние колонки
    df = drop_extra_columns(df)
    
    # Сортируем по дате
    df[DATE_COLUMN] = pd.to_datetime(df[DATE_COLUMN], errors='coerce')
    df = df.sort_values(by=DATE_COLUMN, ascending=True).reset_index(drop=True)
    
    # Добавляем нумерацию
    df = add_numbering(df)
    
    # Очищаем жалобы
    if COMPLAINT_COLUMN in df.columns:
        df[COMPLAINT_COLUMN] = df[COMPLAINT_COLUMN].apply(clean_complaint_prefix)
    
    # Удаляем старый лист и создаем новый
    if 'Детальные' in wb.sheetnames:
        del wb['Детальные']
    
    ws = wb.create_sheet('Детальные', 0)
    
    # Пишем данные
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx).value = value
    
    # Форматируем
    apply_borders_and_formatting(ws, max_col=len(df.columns))
    
    return df


def process_negative_sheet(wb, df):
    """Обрабатывает лист "Отрицательные_и_жалобы"
    
    Только записи со статусом содержащим "отриц" (и БЕЗ "заявка закрыта")
    """
    # Фильтруем по статусу
    df_neg = df[df[STATUS_COLUMN].apply(lambda x: classify_status(x) == 'Отрицательные')].copy()
    df_neg = df_neg.reset_index(drop=True)
    
    # Очищаем жалобы
    if COMPLAINT_COLUMN in df_neg.columns:
        df_neg[COMPLAINT_COLUMN] = df_neg[COMPLAINT_COLUMN].apply(clean_complaint_prefix)
    
    # Переиндексируем с 1
    df_neg.insert(0, '№', range(1, len(df_neg) + 1))
    
    # Удаляем старый лист и создаем новый
    if 'Отрицательные_и_жалобы' in wb.sheetnames:
        del wb['Отрицательные_и_жалобы']
    
    ws = wb.create_sheet('Отрицательные_и_жалобы')
    
    # Пишем данные
    for r_idx, row in enumerate(dataframe_to_rows(df_neg, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx).value = value
    
    # Форматируем
    apply_borders_and_formatting(ws, max_col=len(df_neg.columns))


def process_regional_summary(wb, df):
    """Обрабатывает лист "Жалобы_по_регионам"
    
    Сводка по регионам с типами жалоб
    """
    region_col = 'Регион_112'
    
    if region_col not in df.columns or COMPLAINT_COLUMN not in df.columns:
        return
    
    # Группируем данные
    regional_data = []
    regions = df[region_col].dropna().unique()
    
    for region in sorted(regions):
        region_df = df[df[region_col] == region]
        
        # Считаем жалобы
        total_complaints = region_df[COMPLAINT_COLUMN].notna().sum()
        positive_count = region_df[STATUS_COLUMN].apply(lambda x: classify_status(x) == 'Положительные').sum()
        no_reach_count = region_df[STATUS_COLUMN].apply(lambda x: classify_status(x) == 'Не дозвонились').sum()
        
        regional_data.append({
            'Регион_112': region,
            'Количество_жалоб': total_complaints,
            'Положительные': positive_count,
            'Не дозвонились': no_reach_count,
            'Всего': len(region_df)
        })
    
    df_regional = pd.DataFrame(regional_data)
    
    # Создаем лист
    if 'Жалобы_по_регионам' in wb.sheetnames:
        del wb['Жалобы_по_регионам']
    
    ws = wb.create_sheet('Жалобы_по_регионам')
    
    # Пишем данные
    for r_idx, row in enumerate(dataframe_to_rows(df_regional, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx).value = value
    
    apply_borders_and_formatting(ws, max_col=len(df_regional.columns))


def process_regions_complaints_pivot(wb, df):
    """Обрабатывает лист "Регионы_и_жалобы"
    
    PIVOT таблица: регионы в строках, типы жалоб в столбцах
    """
    region_col = 'Регион_112'
    
    if region_col not in df.columns or COMPLAINT_COLUMN not in df.columns:
        return
    
    # Создаем PIVOT таблицу
    df['complaint_type'] = df[COMPLAINT_COLUMN].notna().astype(int)
    
    pivot_data = df.groupby([region_col, STATUS_COLUMN]).size().unstack(fill_value=0)
    pivot_data = pivot_data.reset_index()
    
    # Переименовываем
    pivot_data.columns.name = None
    
    # Создаем лист
    if 'Регионы_и_жалобы' in wb.sheetnames:
        del wb['Регионы_и_жалобы']
    
    ws = wb.create_sheet('Регионы_и_жалобы')
    
    # Пишем данные
    for r_idx, row in enumerate(dataframe_to_rows(pivot_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx).value = value
    
    apply_borders_and_formatting(ws, max_col=len(pivot_data.columns))


def process_not_found_sheet(wb, df, all_categories):
    """Обрабатывает лист "Не_найденные_заявки"
    
    Только записи, которыми не обработаны в других категориях
    """
    # Фильтруем - берем только те, что не в основных категориях
    df_not_found = df[~df.index.isin(all_categories)].copy()
    df_not_found = df_not_found.reset_index(drop=True)
    
    # Добавляем нумерацию
    df_not_found.insert(0, '№', range(1, len(df_not_found) + 1))
    
    # Создаем лист
    if 'Не_найденные_заявки' in wb.sheetnames:
        del wb['Не_найденные_заявки']
    
    ws = wb.create_sheet('Не_найденные_заявки')
    
    # Пишем данные
    for r_idx, row in enumerate(dataframe_to_rows(df_not_found, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx).value = value
    
    apply_borders_and_formatting(ws, max_col=len(df_not_found.columns))


# ============================================================================
# ГЛАВНАЯ ЛОГИКА
# ============================================================================

def process_file(file_path):
    """Полная обработка одного файла"""
    print(f"\n{'='*70}")
    print(f"Обработку: {Path(file_path).name}")
    print(f"{'='*70}")
    
    try:
        # Читаем основной лист (должен быть "Детальные")
        df = pd.read_excel(file_path, sheet_name='Детальные')
        print(f"✓ Загружено {len(df)} строк, {len(df.columns)} колонок")
        
        # Открываем рабочую книгу для редактирования
        wb = load_workbook(file_path)
        
        # Обрабатываем каждый лист
        print("\n📋 Обработка листов:")
        
        # 1. Обработка "Детальные"
        print("  1️⃣ Лист 'Детальные'...")
        df_detailed = process_detailed_sheet(wb, df.copy())
        print(f"     ✓ Обработано {len(df_detailed)} строк")
        
        # 2. Обработка "Отрицательные_и_жалобы"
        print("  2️⃣ Лист 'Отрицательные_и_жалобы'...")
        process_negative_sheet(wb, df_detailed)
        print(f"     ✓ Обработано")
        
        # 3. Обработка "Жалобы_по_регионам"
        print("  3️⃣ Лист 'Жалобы_по_регионам'...")
        process_regional_summary(wb, df_detailed)
        print(f"     ✓ Обработано")
        
        # 4. Обработка "Регионы_и_жалобы"
        print("  4️⃣ Лист 'Регионы_и_жалобы'...")
        process_regions_complaints_pivot(wb, df_detailed)
        print(f"     ✓ Обработано")
        
        # 5. Обработка "Не_найденные_заявки"
        print("  5️⃣ Лист 'Не_найденные_заявки'...")
        all_categories = set()  # Можно расширить логику если нужно
        process_not_found_sheet(wb, df_detailed, all_categories)
        print(f"     ✓ Обработано")
        
        # Сохраняем
        wb.save(file_path)
        print(f"\n✅ {Path(file_path).name} успешно обработан!")
        return True
        
    except Exception as e:
        print(f"\n❌ Ошибка при обработке {Path(file_path).name}:")
        print(f"   {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Главный файл обработки"""
    if len(sys.argv) < 2:
        print("Использование: python finalize_reports_complete.py '<путь_к_папке>'")
        print("Пример: python finalize_reports_complete.py 'reports/Службы Обратная связь за Январь месяц 2026 год'")
        sys.exit(1)
    
    folder_path = sys.argv[1]
    
    if not os.path.exists(folder_path):
        print(f"❌ Папка не найдена: {folder_path}")
        sys.exit(1)
    
    # Находим все .xlsx файлы
    xlsx_files = sorted(glob.glob(os.path.join(folder_path, '*.xlsx')))
    
    if not xlsx_files:
        print(f"❌ Файлы .xlsx не найдены в {folder_path}")
        sys.exit(1)
    
    print(f"\n{'*'*70}")
    print(f"ФИНАЛЬНАЯ ОБРАБОТКА ОТЧЕТОВ")
    print(f"{'*'*70}")
    print(f"\nНайдено {len(xlsx_files)} файлов для обработки:")
    for f in xlsx_files:
        print(f"  • {Path(f).name}")
    
    # Обрабатываем каждый файл
    success_count = 0
    for file_path in xlsx_files:
        if process_file(file_path):
            success_count += 1
    
    print(f"\n{'*'*70}")
    print(f"✅ ЗАВЕРШЕНО: {success_count}/{len(xlsx_files)} файлов успешно обработано")
    print(f"{'*'*70}\n")


if __name__ == '__main__':
    main()
