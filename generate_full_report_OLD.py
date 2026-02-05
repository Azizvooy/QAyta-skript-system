#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Объединение данных ФИКСА + ЗАЯВКИ и генерация отчетов
"""

import sqlite3
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / 'data' / 'fiksa_database.db'
OUTPUT_DIR = BASE_DIR / 'output' / 'reports'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print('\n' + '='*80)
print('📊 ОБЪЕДИНЕНИЕ ДАННЫХ ФИКСА + ЗАЯВКИ')
print('='*80)

conn = sqlite3.connect(DB_PATH)

# 1. ОБЪЕДИНЕНИЕ ДАННЫХ
print('\n[1/3] Объединение данных из фиксы и заявок...')

query_full = '''
    SELECT 
        fr.id,
        fr.collection_date as "Дата сбора",
        fr.call_date as "Дата звонка",
        DATE(fr.call_date) as "День фиксации",
        fr.operator_name as "Оператор",
        fr.card_number as "Карта",
        COALESCE(fr.phone, ch.caller_phone) as "Телефон",
        fr.full_name as "ФИО",
        fr.address as "Адрес",
        fr.status as "Статус фиксации",
        ch.service_name as "Служба",
        ch.status as "Статус заявки",
        ch.region as "Регион",
        ch.district as "Район",
        ch.incident_number as "Номер инцидента",
        ch.reason as "Причина обращения",
        ch.description as "Описание",
        fr.notes as "Примечания"
    FROM fiksa_records fr
    LEFT JOIN call_history_112 ch ON fr.card_number = ch.card_number
    WHERE fr.operator_name IS NOT NULL
    ORDER BY fr.call_date DESC, fr.operator_name
'''

df_full = pd.read_sql_query(query_full, conn)
print(f'  Загружено: {len(df_full):,} записей')
print(f'  С заполненным телефоном: {df_full["Телефон"].notna().sum():,}')
print(f'  Со службой: {df_full["Служба"].notna().sum():,}')

# 2. ОТЧЕТ ПО ДНЯМ ФИКСАЦИИ
print('\n[2/3] Генерация отчета по дням фиксации...')

query_daily = '''
    SELECT 
        DATE(call_date) as "Дата",
        COUNT(*) as "Всего фиксаций",
        COUNT(DISTINCT operator_name) as "Операторов работало",
        COUNT(CASE WHEN status LIKE '%Положительн%' THEN 1 END) as "✓ Положительных",
        COUNT(CASE WHEN status LIKE '%Отрицательн%' THEN 1 END) as "✗ Отрицательных",
        COUNT(CASE WHEN status LIKE '%Недозвон%' THEN 1 END) as "⚠ Недозвонились",
        ROUND(COUNT(CASE WHEN status LIKE '%Положительн%' THEN 1 END) * 100.0 / COUNT(*), 1) as "% Положит.",
        ROUND(COUNT(CASE WHEN status LIKE '%Отрицательн%' THEN 1 END) * 100.0 / COUNT(*), 1) as "% Отрицат.",
        ROUND(COUNT(*) * 1.0 / COUNT(DISTINCT operator_name), 1) as "Среднее на оператора"
    FROM fiksa_records
    WHERE call_date IS NOT NULL AND operator_name IS NOT NULL
    GROUP BY DATE(call_date)
    ORDER BY DATE(call_date) DESC
    LIMIT 60
'''

df_daily = pd.read_sql_query(query_daily, conn)
print(f'  Сгенерировано: {len(df_daily)} дней')

# 3. РЕЙТИНГ ОПЕРАТОРОВ
print('\n[3/3] Генерация рейтинга операторов...')

query_rating = '''
    SELECT 
        operator_name as "Оператор",
        COUNT(*) as "Всего фиксаций",
        COUNT(CASE WHEN status LIKE '%Положительн%' THEN 1 END) as "✓ Положительных",
        COUNT(CASE WHEN status LIKE '%Отрицательн%' THEN 1 END) as "✗ Отрицательных",
        COUNT(CASE WHEN status LIKE '%Недозвон%' THEN 1 END) as "⚠ Недозвонились",
        COUNT(CASE WHEN status LIKE '%Нет ответа%' OR status LIKE '%занято%' THEN 1 END) as "⊗ Нет ответа",
        ROUND(COUNT(CASE WHEN status LIKE '%Положительн%' THEN 1 END) * 100.0 / COUNT(*), 1) as "% Положит.",
        ROUND(COUNT(CASE WHEN status LIKE '%Отрицательн%' THEN 1 END) * 100.0 / COUNT(*), 1) as "% Отрицат.",
        ROUND(
            (COUNT(CASE WHEN status LIKE '%Положительн%' THEN 1 END) * 100.0) / 
            NULLIF(COUNT(CASE WHEN status LIKE '%Положительн%' OR status LIKE '%Отрицательн%' THEN 1 END), 0), 
            1
        ) as "🏆 Рейтинг",
        COUNT(DISTINCT DATE(call_date)) as "Рабочих дней",
        ROUND(COUNT(*) * 1.0 / NULLIF(COUNT(DISTINCT DATE(call_date)), 0), 1) as "Среднее/день"
    FROM fiksa_records
    WHERE operator_name IS NOT NULL AND status IS NOT NULL
    GROUP BY operator_name
    ORDER BY COUNT(*) DESC
'''

df_rating = pd.read_sql_query(query_rating, conn)
print(f'  Операторов в рейтинге: {len(df_rating)}')

conn.close()

# СОХРАНЕНИЕ В EXCEL
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = OUTPUT_DIR / f'📊_ПОЛНЫЙ_ОТЧЕТ_{timestamp}.xlsx'

print(f'\n[СОХРАНЕНИЕ] {output_file.name}')

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Лист 1: Полные данные
    df_full.to_excel(writer, sheet_name='Полные данные', index=False)
    
    # Лист 2: По дням
    df_daily.to_excel(writer, sheet_name='📅 По дням фиксации', index=False)
    
    # Лист 3: Рейтинг
    df_rating.to_excel(writer, sheet_name='🏆 Рейтинг операторов', index=False)

# Форматирование
wb = load_workbook(output_file)

# Стили
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
positive_font = Font(name='Calibri', size=10, color='006100')
negative_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
negative_font = Font(name='Calibri', size=10, color='9C0006')
warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
warning_font = Font(name='Calibri', size=10, color='9C5700')
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

def format_sheet(ws):
    # Заголовки
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Автоширина
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
    
    # Заморозка заголовков
    ws.freeze_panes = 'A2'
    
    # Автофильтр
    ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

# Форматирование всех листов
for sheet_name in wb.sheetnames:
    format_sheet(wb[sheet_name])

# Условное форматирование для рейтинга
ws_rating = wb['🏆 Рейтинг операторов']
for row_idx, row in enumerate(ws_rating.iter_rows(min_row=2, max_row=ws_rating.max_row), start=2):
    # Колонка "🏆 Рейтинг" (колонка 9)
    rating_cell = row[8]
    try:
        rating_val = float(rating_cell.value) if rating_cell.value else 0
        if rating_val >= 80:
            rating_cell.fill = positive_fill
            rating_cell.font = positive_font
        elif rating_val >= 60:
            rating_cell.fill = warning_fill
            rating_cell.font = warning_font
        elif rating_val > 0:
            rating_cell.fill = negative_fill
            rating_cell.font = negative_font
    except:
        pass

# Условное форматирование для дневного отчета
ws_daily = wb['📅 По дням фиксации']
for row_idx, row in enumerate(ws_daily.iter_rows(min_row=2, max_row=ws_daily.max_row), start=2):
    # Колонка "% Положит." (колонка 7)
    percent_cell = row[6]
    try:
        percent_val = float(percent_cell.value) if percent_cell.value else 0
        if percent_val >= 30:
            percent_cell.fill = positive_fill
            percent_cell.font = positive_font
        elif percent_val >= 20:
            percent_cell.fill = warning_fill
            percent_cell.font = warning_font
        elif percent_val > 0:
            percent_cell.fill = negative_fill
            percent_cell.font = negative_font
    except:
        pass

wb.save(output_file)

# СТАТИСТИКА
print('\n' + '='*80)
print('✅ ОТЧЕТ ГОТОВ')
print('='*80)
print(f'\n📊 Файл: {output_file.name}')
print(f'📈 Размер: {output_file.stat().st_size / 1024:.1f} КБ')
print(f'\n📋 Листы:')
print(f'  1. Полные данные - {len(df_full):,} записей')
print(f'  2. По дням фиксации - {len(df_daily)} дней')
print(f'  3. Рейтинг операторов - {len(df_rating)} операторов')

# Топ-5 дней
print(f'\n📅 ТОП-5 ДНЕЙ ПО ФИКСАЦИЯМ:')
for idx, row in df_daily.head(5).iterrows():
    print(f'  {row["Дата"]}: {int(row["Всего фиксаций"]):,} фиксаций ({row["% Положит."]}% положит.)')

# Топ-5 операторов
print(f'\n🏆 ТОП-5 ОПЕРАТОРОВ:')
for idx, row in df_rating.head(5).iterrows():
    rating = row["🏆 Рейтинг"] if pd.notna(row["🏆 Рейтинг"]) else 0
    print(f'  {row["Оператор"][:40]:40} | {int(row["Всего фиксаций"]):6,} фикс. | Рейтинг: {rating:.1f}%')

print('\n' + '='*80)
