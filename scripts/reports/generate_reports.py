#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
=================================================================
ГЕНЕРАЦИЯ EXCEL ОТЧЕТОВ С ОТРИЦАТЕЛЬНЫМИ ОТЗЫВАМИ
=================================================================
Скрипт экспортирует отрицательные отзывы в Excel файлы.

Структура отчетов:
- ТОЛЬКО_ОТРИЦАТЕЛЬНЫЕ/
  ├── СЛУЖБА_101_ОТРИЦАТЕЛЬНЫЕ.xlsx
  ├── СЛУЖБА_102_ОТРИЦАТЕЛЬНЫЕ.xlsx
  └── ОБЩИЙ_ОТЧЕТ_ОТРИЦАТЕЛЬНЫЕ.xlsx

Особенности:
- Группирует по видам услуг (101, 102, 103, 104...)
- Форматирует Excel файлы (шапка, цвета)
- Сохраняет в папку reports/{МЕСЯЦ_ГОД}/ТОЛЬКО_ОТРИЦАТЕЛЬНЫЕ/

Использование:
    python scripts/reports/generate_reports.py

Конфигурация: config/postgresql.env
=================================================================
"""

import os
import psycopg2
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Загрузка конфигурации
CONFIG_DIR = Path(__file__).parent.parent.parent / 'config'
load_dotenv(CONFIG_DIR / 'postgresql.env')

# Параметры подключения к БД
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': os.getenv('DB_PORT', '5432'),
    'database': os.getenv('DB_NAME', 'qayta_data'),
    'user': os.getenv('DB_USER', 'qayta_user'),
    'password': os.getenv('DB_PASSWORD', 'qayta_password_2026')
}


def get_db_connection():
    """Создать подключение к БД"""
    return psycopg2.connect(**DB_CONFIG)


def format_excel(file_path, sheet_name='Sheet1'):
    """Отформатировать Excel файл"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Стиль заголовка (синий фон, белый текст)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Отформатировать первую строку
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Установить ширину колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(file_path)


def main():
    """Генерировать отчеты"""
    print("="*70)
    print("ГЕНЕРАЦИЯ EXCEL ОТЧЕТОВ С ОТРИЦАТЕЛЬНЫМИ ОТЗЫВАМИ")
    print("="*70)
    
    conn = get_db_connection()
    start = datetime.now()
    
    # Определяем папку для сохранения
    month_dir = start.strftime('%Y-%m')
    reports_base = Path(__file__).parent.parent.parent / 'reports' / month_dir
    reports_dir = reports_base / 'ТОЛЬКО_ОТРИЦАТЕЛЬНЫЕ'
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\n📁 Папка для отчетов: {reports_dir}\n")
    
    cursor = conn.cursor()
    
    # 1. Загрузить все отрицательные отзывы
    print("1️⃣  Загрузка отрицательных отзывов из БД...")
    df_all = pd.read_sql(
        "SELECT * FROM negative_complaints ORDER BY call_datetime DESC",
        conn
    )
    
    if len(df_all) == 0:
        print("⚠️  Нет отрицательных отзывов в БД")
        cursor.close()
        conn.close()
        return
    
    total_records = len(df_all)
    print(f"✓ Загружено {total_records:,} отрицательных отзывов\n")
    
    # 2. Генерировать отчеты по услугам
    print("2️⃣  Генерирование отчетов по видам услуг...")
    
    services = df_all['service'].dropna().unique()
    if None in df_all['service'].values:
        services = list(services) + [None]
    
    service_reports = {}
    
    for service in sorted([s for s in services if s]):
        df_service = df_all[df_all['service'] == service]
        
        if len(df_service) > 0:
            # Извлекаем номер услуги (101, 102, 103...)
            service_num = service[:3] if len(service) >= 3 else service
            
            filename = f"СЛУЖБА_{service_num}_ОТРИЦАТЕЛЬНЫЕ.xlsx"
            file_path = reports_dir / filename
            
            # Сохраняем в Excel
            df_service.to_excel(file_path, sheet_name='Отрицательные', index=False)
            format_excel(file_path, 'Отрицательные')
            
            service_reports[service] = len(df_service)
            print(f"   ✓ {filename}: {len(df_service):,} записей")
    
    # 3. Общий отчет все отрицательные
    print("\n3️⃣  Генерирование общего отчета...")
    
    general_file = reports_dir / "ОБЩИЙ_ОТЧЕТ_ОТРИЦАТЕЛЬНЫЕ.xlsx"
    df_all.to_excel(general_file, sheet_name='Отрицательные', index=False)
    format_excel(general_file, 'Отрицательные')
    print(f"   ✓ ОБЩИЙ_ОТЧЕТ_ОТРИЦАТЕЛЬНЫЕ.xlsx: {total_records:,} записей")
    
    # 4. Статистика
    print("\n4️⃣  Генерирование сводной статистики...")
    
    stats_data = {
        'Вид услуги': list(service_reports.keys()),
        'Количество отрицательных': list(service_reports.values())
    }
    df_stats = pd.DataFrame(stats_data)
    df_stats['Процент'] = (df_stats['Количество отрицательных'] / total_records * 100).round(2)
    df_stats = df_stats.sort_values('Количество отрицательных', ascending=False)
    
    stats_file = reports_dir / "СТАТИСТИКА_ОТРИЦАТЕЛЬНЫЕ.xlsx"
    df_stats.to_excel(stats_file, sheet_name='Статистика', index=False)
    format_excel(stats_file, 'Статистика')
    print(f"   ✓ СТАТИСТИКА_ОТРИЦАТЕЛЬНЫЕ.xlsx создана")
    
    cursor.close()
    conn.close()
    
    # Финальная информация
    elapsed = datetime.now() - start
    print("\n" + "="*70)
    print("✅ ОТЧЕТЫ УСПЕШНО СОЗДАНЫ")
    print("="*70)
    print(f"Время выполнения: {elapsed}")
    print(f"\n📊 Созданные файлы:")
    
    for f in sorted(reports_dir.glob("*.xlsx")):
        size = f.stat().st_size / 1024
        print(f"   • {f.name} ({size:.1f} KB)")
    
    print(f"\n📁 Путь к отчетам: {reports_dir}")
    print("="*70 + "\n")


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n❌ ОШИБКА: {e}\n")
        import traceback
        traceback.print_exc()
