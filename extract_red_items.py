"""
Извлечение всех строк с красным фоном из всех отчетов
Это строки, которые нужно исключить из итоговых отчетов
"""
import openpyxl
from pathlib import Path
import pandas as pd

examples_dir = Path('ПРИМЕРЫ_ОТЧЕТОВ')

print('=' * 80)
print('СТРОКИ ДЛЯ УДАЛЕНИЯ (отмечены КРАСНЫМ)')
print('=' * 80)
print()

all_red_items = {}

for file_path in sorted(examples_dir.glob('*.xlsx')):
    print(f'\n📄 {file_path.name}')
    print('-' * 80)
    
    try:
        wb = openpyxl.load_workbook(file_path)
        file_red_items = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Ищем строки с красным фоном
            red_rows = set()
            
            for row in ws.iter_rows(min_row=1, max_row=100):
                for cell in row:
                    if not cell.fill or not cell.fill.fgColor:
                        continue
                    
                    if not hasattr(cell.fill.fgColor, 'rgb') or not cell.fill.fgColor.rgb:
                        continue
                    
                    rgb = str(cell.fill.fgColor.rgb)
                    if 'FF0000' in rgb or rgb.startswith('FFFF0000'):
                        red_rows.add(cell.row)
            
            # Извлекаем данные из красных строк
            for row_num in sorted(red_rows):
                row_data = []
                for cell in ws[row_num]:
                    if cell.value:
                        row_data.append(str(cell.value))
                
                if row_data:
                    item_text = ' | '.join(row_data[:5])  # Первые 5 столбцов
                    file_red_items.append({
                        'sheet': sheet_name,
                        'row': row_num,
                        'data': item_text,
                        'first_value': row_data[0] if row_data else ''
                    })
                    print(f'  Строка {row_num} [{sheet_name}]: {item_text}')
        
        if not file_red_items:
            print('  ✅ Нет красных строк')
        
        all_red_items[file_path.name] = file_red_items
        
    except Exception as e:
        print(f'  ❌ Ошибка: {e}')

# Создаем список паттернов для фильтрации
print('\n' + '=' * 80)
print('ПАТТЕРНЫ ДЛЯ ИСКЛЮЧЕНИЯ:')
print('=' * 80)

exclude_patterns = set()
for file_name, items in all_red_items.items():
    for item in items:
        first_val = item['first_value'].strip()
        if first_val and first_val not in ['0', '']:
            exclude_patterns.add(first_val)

if exclude_patterns:
    print('\nСтроки, содержащие в первом столбце:')
    for pattern in sorted(exclude_patterns):
        print(f'  • "{pattern}"')
    
    # Сохраняем в файл
    with open('СТРОКИ_ДЛЯ_УДАЛЕНИЯ.txt', 'w', encoding='utf-8') as f:
        f.write('# Паттерны для исключения из отчетов\n')
        f.write('# Эти строки отмечены красным и должны быть удалены\n\n')
        for pattern in sorted(exclude_patterns):
            f.write(f'{pattern}\n')
    
    print('\n💾 Сохранено в СТРОКИ_ДЛЯ_УДАЛЕНИЯ.txt')
else:
    print('\n✅ Не найдено паттернов для исключения')

print('\n' + '=' * 80)
